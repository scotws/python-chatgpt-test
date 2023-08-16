import os
import subprocess
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from io import BytesIO

THUMBNAIL_WIDTH = 200  # Adjust this value as needed

def get_metadata(file_path):
    cmd = ["identify", "-format", "'%[parameters]'", file_path]
    result = subprocess.check_output(cmd).decode('utf-8')
    artist = result.split(":")[0].split("(")[1].strip()
    prompt = result.split(")")[1].split("Negative")[0].strip()
    negative_prompt = result.split("Negative prompt:")[1].split("Steps:")[0].strip()
    model = result.split("Model:")[1].split(",")[0].strip()

    return artist, prompt, negative_prompt, model

def create_excel(data, folder_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stable Diffusion Benchmark"

    # Set title and model
    ws[f"A1"] = f"Stable Diffusion Benchmark {data['date']}"
    ws[f"A2"] = f"Model: {data['model']}"

    # Set headers
    headers = ["Artist"] + list(data['prompts'])
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}4"]
        cell.value = header
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # Wrap the text for promptsk

    # Adjust the height of the header row
    ws.row_dimensions[4].height = 30  # Adjust this value as needed

    # Set column widths
    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = THUMBNAIL_WIDTH / 6  # Approximate width in Excel units

    # Populate data
    row_num = 5

    for artist in sorted(data['artists'].keys()):
        prompts = data['artists'][artist]
        cell = ws[f"A{row_num}"]
        cell.value = artist
        cell.alignment = openpyxl.styles.Alignment(vertical="center")  # Center artist names vertically

        for col_num, prompt in enumerate(data['prompts'], 2):
            if prompt in prompts:
                img_path = os.path.join(folder_path, prompts[prompt])
                with PILImage.open(img_path) as img:
                    aspect_ratio = img.height / img.width
                    new_height = int(THUMBNAIL_WIDTH * aspect_ratio)
                    img.thumbnail((THUMBNAIL_WIDTH, new_height))
                    img_byte_array = BytesIO()
                    img.save(img_byte_array, format='PNG')
                    img_excel = ExcelImage(img_byte_array)
                    ws.add_image(img_excel, f"{get_column_letter(col_num)}{row_num}")
        # Adjust row height
        ws.row_dimensions[row_num].height = new_height * 0.75  # Approximate height in Excel units
        row_num += 1

    # Save the workbook
    wb.save(os.path.join(folder_path, "output.xlsx"))


def main():
    folder_path = input("Enter the folder location (or press Enter for current directory): ")
    if not folder_path:
        folder_path = os.getcwd()

    if not os.access(folder_path, os.R_OK | os.W_OK):
        print("Error: No read/write permissions for the folder.")
        return

    if not os.path.isfile("/opt/homebrew/bin/identify"):
        print("Error: 'identify' program is not installed or not executable.")
        return

    png_files = [f for f in os.listdir(folder_path) if f.endswith('.png')]
    if not png_files:
        print("Error: No PNG files found in the folder.")
        return

    data = {
        "date": os.path.basename(folder_path),
        "model": "",
        "negative_prompt": "",
        "prompts": [],
        "artists": {}
    }

    for png_file in png_files:
        artist, prompt, negative_prompt, model = get_metadata(os.path.join(folder_path, png_file))

        if not data["model"]:
            data["model"] = model
        if not data["negative_prompt"]:
            data["negative_prompt"] = negative_prompt

        if prompt not in data["prompts"]:
            data["prompts"].append(prompt)

        if artist not in data["artists"]:
            data["artists"][artist] = {}

        data["artists"][artist][prompt] = png_file

    create_excel(data, folder_path)

if __name__ == "__main__":
    main()

